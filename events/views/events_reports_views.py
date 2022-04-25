import datetime

import openpyxl
from openpyxl.utils import get_column_letter

from django.http import HttpResponse, HttpResponseRedirect
from django.views.generic import View
from django.shortcuts import render
from django.db.models import Q, F, Sum, Prefetch
from django.contrib import messages

from store.models import Purchase
from myapa.viewmixins import AuthenticateProviderContactRoleMixin
from registrations.models import Attendee

from events.models import Event


class EventReports(AuthenticateProviderContactRoleMixin, View):

    template_name = "events/newtheme/exports/attendees-export.html"
    context = {}
    content = None

    def set_content(self, request, *args, **kwargs):
        self.content = Event.objects.get(master__id=kwargs["master_id"], publish_status="PUBLISHED")

    def dispatch(self, request, *args, **kwargs):
        self.set_content(request, *args, **kwargs)
        self.context["event"] = self.content
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):

        return self.render_template(request)

    def render_template(self, request):
        return render(request, self.template_name, self.context)

class ReportsDownload(AuthenticateProviderContactRoleMixin, View):
    context = {}
    content = None
    master_id = None
    def set_content(self, request, *args, **kwargs):
        self.content = Event.objects.get(master__id=kwargs["master_id"], publish_status="PUBLISHED")
        self.master_id = self.content.master.id

    def get(self, request, *args, **kwargs):
        return self.render_template(request, *args, **kwargs)

    def render_template(self, request, *args, **kwargs):
        return self.export_reports(request, *args, **kwargs)

    def export_reports(self, request, *args, **kwargs):
        request_options = request.GET
        # if request_options['from_date'] and request_options['to_date']:
        report_type = request_options['report_options']
        begin_time = request_options["from_date"]
        end_time = request_options["to_date"]

        if begin_time:
            begin_time = datetime.datetime.strptime(request_options['from_date'], "%Y-%m-%d")
        if end_time:
            end_time = datetime.datetime.strptime(request_options['to_date'], "%Y-%m-%d").replace(hour = 23, minute = 59, second = 59)
            #end_time = end_time + datetime.timedelta(days=1)

        if report_type == "registration":
            return attendees_event_registrations(request, kwargs["master_id"], report_type, begin_time, end_time)
        elif report_type =="activity_registration":
            return attendees_activity_registrations(request, kwargs["master_id"], report_type, begin_time, end_time)
        elif report_type =="and_activity_registration":
            return attendees_event_activity_registrations(request, kwargs["master_id"], report_type, begin_time, end_time)
        elif report_type =="activitysummaryreport":
            return activitysummaryreport(request, kwargs["master_id"], report_type, begin_time, end_time)
        elif report_type =="financialsummaryreport":
            return financialsummaryreport(request, kwargs["master_id"], report_type, begin_time, end_time)
        elif report_type == "cancelledreport":
            return cancellationreport(request, kwargs["master_id"], report_type, begin_time, end_time)
        elif report_type == "all_purchase_report":
            return all_purchase_report(request, kwargs["master_id"], report_type, begin_time, end_time)
        elif report_type == "event_option_report":
            return event_option_report(request, kwargs["master_id"], report_type, begin_time, end_time)
        # elif report_type == "registration_summary":
        #     from_date = datetime.datetime.strptime(request_options['from_date'], "%Y-%m-%d")
        #     to_date = datetime.datetime.strptime(request_options['to_date'], "%Y-%m-%d").replace(hour = 23, minute = 59, second = 59)
        #     return registration_summary_report(request, kwargs["master_id"], report_type)
        else:
            messages.error(request, "Please select a report type.")
            return HttpResponseRedirect("/events/manage/{0}/reports/".format(str(self.master_id)))

        # else:
        #     messages.error(request, "Please select start and end dates to filter on.")
        #     return HttpResponseRedirect("/events/manage/{0}/reports/".format(str(self.master_id)))


def convert_str_datetime(str_dt):
    return datetime.datetime.strptime(str_dt, "%Y-%m-%d")


def attendees_event_registrations(request, master_id, report_type, begin_time, end_time):
    """
    returns a report for all registrations for this event with the dates submitted
    """
    #return HttpResponseRedirect("/events/manage/{0}/reports/".format(str(master_id)))
    request_options = request.GET
    filter_kwargs = { "event__master__id":master_id, 
                    "purchase__status":"A", "status":"A",} 
    if begin_time and end_time:
        filter_kwargs["purchase__order__submitted_time__range"] = (begin_time, end_time)
    elif begin_time:
        filter_kwargs["purchase__order__submitted_time__gt"] = begin_time
    elif end_time:
        filter_kwargs["purchase__order__submitted_time__lt"] = end_time


    attendees = Attendee.objects.filter(**filter_kwargs).select_related("purchase").select_related("contact__user").order_by("purchase__order__submitted_time")

    response = HttpResponse(content_type="application/vnd.openxml")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=Event_" + report_type + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Event_" + report_type

    if attendees:
        row_num = 0

        question_1 = "Question 1 (not used)"
        question_2 = "Question 2 (not used)"
        question_3 = "Question 3 (not used)"

        agreement_statement_1 = "Agreement Response 1 (not used)"
        agreement_statement_2 = "Agreement Response 2 (not used)"
        agreement_statement_3 = "Agreement Response 3 (not used)"

        if hasattr(attendees[0].event.product, "question_1"):
            question_1 = attendees[0].event.product.question_1
            if question_1 == "" or question_1 == None:
                question_1 = "Question 1 (not used)"

        if hasattr(attendees[0].event.product, "question_2"):
            question_2 = attendees[0].event.product.question_2
            if question_2 == "":
                question_2 = "Question 2 (not used)"

        if hasattr(attendees[0].event.product, "question_3"):
            question_3 = attendees[0].event.product.question_3
            if question_3 == "":
                question_3 = "Question 3 (not used)"
        
        if hasattr(attendees[0].event.product, "agreement_statement_1"):
            agreement_statement_1 = attendees[0].event.product.agreement_statement_1
            if agreement_statement_1 == "":
                agreement_statement_1 = "Agreement Response 1 (not used)"

        if hasattr(attendees[0].event.product, "agreement_statement_2"):
            agreement_statement_2 = attendees[0].event.product.agreement_statement_2
            if agreement_statement_2 == "":
                agreement_statement_2 = "Agreement Response 2 (not used)"
        
        if hasattr(attendees[0].event.product, "agreement_statement_3"):
            agreement_statement_3 = attendees[0].event.product.agreement_statement_3
            if agreement_statement_3 == "":
                agreement_statement_3 = "Agreement Response 3 (not used)"

        columns = [
                    ("Master ID",20), 
                    #("Activity Name",20), 
                    #("Activity Code",20),
                    ("Registration option", 20), 
                    ("Price Title", 50),
                    ("WebUserID", 10),
                    ("First Name",20), 
                    ("Middle Name",20), 
                    ("Last Name",20), 
                    ("Suffix", 10),
                    ("Designation",20),
                    (question_1, 50),
                    (question_2, 50),
                    (question_3, 50),
                    (agreement_statement_1, 50),
                    (agreement_statement_2, 50),
                    (agreement_statement_3, 50),

                    ("Member Type",20), 
                    ("Company",20), 
                    ("Address 1",20), 
                    ("Address 2",20), 
                    ("City",20), 
                    ("State",20), 
                    ("Zip Code", 10),
                    ("Country", 20), 
                    ("Email", 20),
                    ("Phone",10), 
                    ("Date Purchased", 20),]


        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for attendee in attendees:
            contact = attendee.contact

            row_num += 1

            question_response_1 = ''
            question_response_2 = ''
            question_response_3 = ''

            if hasattr(attendee.purchase, "question_response_1"):
                question_response_1 = attendee.purchase.question_response_1
                if question_response_1 == None:
                    question_response_1 = ''
            if hasattr(attendee.purchase, "question_response_2"):
                question_response_2 = attendee.purchase.question_response_2
                if question_response_2 == None:
                    question_response_2 = ''
            if hasattr(attendee.purchase, "question_response_3"):
                question_response_3 = attendee.purchase.question_response_3
                if question_response_3 == None:
                    question_response_3 = ''
            

            agreement_response_1 = ''
            agreement_response_2 = ''
            agreement_response_3 = ''
            if hasattr(attendee.purchase, "agreement_response_1"):
                agreement_response_1 = attendee.purchase.agreement_response_1
                if agreement_response_1 == None:
                    question_response_1 = ''
            if hasattr(attendee.purchase, "agreement_response_2"):
                agreement_response_2 = attendee.purchase.agreement_response_2
                if agreement_response_2 == None:
                    agreement_response_2 = ''
            if hasattr(attendee.purchase, "agreement_response_3"):
                agreement_response_3 = attendee.purchase.agreement_response_3
                if agreement_response_3 == None:
                    agreement_response_3 = ''
                     
            row = [
                master_id,
                #attendee.event.title,
                #attendee.purchase.product.code,
                attendee.purchase.option,
                attendee.purchase.product_price.title,
                contact.user.username,
                contact.first_name,
                contact.middle_name,
                contact.last_name,
                contact.suffix_name,
                contact.designation,
                question_response_1,
                question_response_2,
                question_response_3,
                agreement_response_1,
                agreement_response_2,
                agreement_response_3,
                contact.member_type, 
                #contact.title,
                contact.company,
                contact.address1,
                contact.address2,
                contact.city,
                contact.state,
                contact.zip_code,
                contact.country,
                contact.email,
                contact.phone,
                None if not attendee.purchase.submitted_time else attendee.purchase.submitted_time.date(),
                ]
                # if event.event_type == ACTIVITY --> Attendee.objects.filter(event = event).count()
                # if event.event_type != ACTIVITY --> Attendee.objects.filter(event = event).count()
           
            for col_num in range(len(row)):
                c = ws.cell(row=row_num+1, column=col_num+1)
                try:
                    c.value = float(row[col_num])
                except:
                    c.value = str(row[col_num])

        wb.save(response)
        return response

    else:
        messages.error(request, "No records were found with the dates passed in. Please change the begin and end date filters and try again.")
        return HttpResponseRedirect("/events/manage/{0}/reports/".format(str(master_id)))


def attendees_activity_registrations(request, master_id, report_type, begin_time, end_time):
    """
    returns all activities for the event based on the dates passed in
    """

    request_options = request.GET
    filter_kwargs = { "parent__id":master_id, "publish_status":"PUBLISHED" } 
    attendee_filter_kwargs = {}
    attendee_filter_kwargs["status"] = "A"
    attendee_filter_kwargs["purchase__status"] = "A"
    attendee_filter_kwargs_no_product = {}

    attendee_filter_kwargs_no_product["purchase__isnull"] = True
    attendee_filter_kwargs_no_product["event__product__isnull"] = True
    attendee_filter_kwargs_no_product["status"] = "A"

    if begin_time and end_time:
        attendee_filter_kwargs_no_product["added_time__range"] = (begin_time, end_time)
        attendee_filter_kwargs["purchase__order__submitted_time__range"] = (begin_time, end_time)
    elif begin_time:
        attendee_filter_kwargs_no_product["added_time__gt"] = begin_time
        attendee_filter_kwargs["purchase__order__submitted_time__gt"] = begin_time
    elif end_time:
        attendee_filter_kwargs_no_product["added_time__lt"] = end_time
        attendee_filter_kwargs["purchase__order__submitted_time__lt"] = end_time

    # if hasattr(event, "product"):
    #     attendees = Attendee.objects.filter(**attendee_filter_kwargs).exclude(purchase__order__isnull=True).select_related("contact").select_related("purchase").order_by("purchase__order__submitted_time")
    # else:
    #     attendees = Attendee.objects.filter(**attendee_filter_kwargs_no_product).select_related("contact").order_by("added_time")

    events = Event.objects.filter(**filter_kwargs).select_related("product").prefetch_related(
        Prefetch("attendees", queryset=Attendee.objects.filter( Q(**attendee_filter_kwargs) | Q(**attendee_filter_kwargs_no_product), status="A").select_related(
            "purchase__order", "purchase__product", "contact", "contact__user"
        ).order_by(
            "purchase__order__submitted_time", "added_time"
        ), to_attr="conf_attendees")
    ).order_by("master__id")

    response = HttpResponse(content_type="application/vnd.openxml")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=Event_" + report_type + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Event_" + report_type

    row_num = 0


    columns = [
        ("Activity ID",20), 
        ("Activity Name",20), 
        #("Activity Code",20),
        ("WebUserID",10),
        ("First Name",20), 
        ("Middle Name",20), 
        ("Last Name",20), 
        ("Suffix", 10),
        #("iMIS Code",15),
        #("Registration option", 20), 
        ("Designation",20),
        ("Member Type",20),
        #("Attendee Title", 20), 
        ("Company",20), 
        ("Address 1",20), 
        ("Address 2",20), 
        ("City",20), 
        ("State",20), 
        ("Zip Code", 10),
        ("Country", 20), 
        ("Email", 20),
        ("Phone",10), 
        ("Purchase Date", 20),
        #("Quantity", 10),
        ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for event in events:
        attendee_filter_kwargs["event__id"] = event.id
        
        # if hasattr(event, "product"):
        #     attendees = Attendee.objects.filter(**attendee_filter_kwargs).exclude(purchase__order__isnull=True).select_related("contact").select_related("purchase").order_by("purchase__order__submitted_time")
        # else:
        #     attendees = Attendee.objects.filter(**attendee_filter_kwargs_no_product).select_related("contact").order_by("added_time")
        # for attendee in attendees:
        for attendee in event.conf_attendees:
            submitted_time = None
            purchase_quantity = None

            if hasattr(event, "product"):
                submitted_time = attendee.purchase.order.submitted_time
                #purchase_quantity = attendee.purchase.quantity
            else:
                submitted_time = attendee.added_time
                #purchase_quantity = 1


            # if row_num == 0:
            #     c1 = ws.cell(row = 1, column = 8)
            #     c2 = ws.cell(row = 1, column = 9)
            #     c3 = ws.cell(row = 1, column = 10)
        
            #     if hasattr(attendee.purchase, "product"):
            #         question_1 = attendee.purchase.product.question_1
            #         if question_1 == "":
            #             question_1 = "Question 1"
            #         c1.value = question_1
            #         question_2 = attendee.purchase.product.question_2
            #         if question_2 == "":
            #             question_2 = "Question 2"
            #         c2.value = question_2
            #         question_3 = attendee.purchase.product.question_3
            #         if question_3 == "":
            #             question_3 = "Question 3"
            #         c3.value = question_3
            #     else:
            #         c1.value = 'Question 1'
            #         c2.value = 'Question 2'
            #         c3.value = 'Question 3'



            row_num += 1
            row = [
                event.master.id,
                event.title,
                #attendee.purchase.code,
                attendee.contact.user.username,
                attendee.contact.first_name,
                attendee.contact.middle_name,
                attendee.contact.last_name,
                attendee.contact.suffix_name,
                #purchase.product.imis_code,
                #purchase.option,
                attendee.contact.designation,
                attendee.contact.member_type,
                #contact.member_title,
                attendee.contact.company,
                attendee.contact.address1,
                attendee.contact.address2,
                attendee.contact.city,
                attendee.contact.state,
                attendee.contact.zip_code,
                attendee.contact.country,
                attendee.contact.email,
                attendee.contact.phone,
                submitted_time.date(),
                #purchase_quantity, #Attendee.objects.filter(event = event, contact=attendee.contact).count() if event.event_type == "ACTIVITY" else 0,
                ]

            for col_num in range(len(row)):
                c = ws.cell(row=row_num+1, column=col_num+1)
                try:
                    c.value = float(row[col_num])
                except:
                    c.value = str(row[col_num])

    wb.save(response)
    return response
    

def attendees_event_activity_registrations(request, master_id, report_type, begin_time, end_time):
    """
    add purchase option code?
    how to filter by dates for this??? not sure this is possible since not all activities have purchases. which date should i use?
    """
    request_options = request.GET
    filter_kwargs = {"publish_status":"PUBLISHED" } 
    attendee_filter_kwargs_master = {}
    attendee_filter_kwargs_master["status"] = "A"
    attendee_filter_kwargs_master["event__master__id"] = master_id
    attendee_filter_kwargs_master["purchase__status"] = "A"

    attendee_filter_kwargs_children_no_product = {}
    attendee_filter_kwargs_children_no_product["purchase__isnull"] = True
    attendee_filter_kwargs_children_no_product["event__product__isnull"] = True
    attendee_filter_kwargs_children_no_product["status"] = "A"
    attendee_filter_kwargs_children_no_product["event__parent__id"] = master_id

    attendee_filter_kwargs_children_product = {}
    attendee_filter_kwargs_children_product["status"] = "A"
    attendee_filter_kwargs_children_product["purchase__status"] = "A"
    attendee_filter_kwargs_children_product["event__parent__id"] = master_id

    if begin_time and end_time:
        attendee_filter_kwargs_children_no_product["added_time__range"] = (begin_time, end_time)
        attendee_filter_kwargs_master["purchase__order__submitted_time__range"] = (begin_time, end_time)
    elif begin_time:
        attendee_filter_kwargs_children_no_product["added_time__gt"] = begin_time
        attendee_filter_kwargs_master["purchase__order__submitted_time__gt"] = begin_time
    elif end_time:
        attendee_filter_kwargs_children_no_product["added_time__lt"] = end_time
        attendee_filter_kwargs_master["purchase__order__submitted_time__lt"] = end_time


    events = Event.objects.filter(Q(parent__id=master_id) | Q(master__id=master_id), publish_status="PUBLISHED").select_related("product").prefetch_related(
        Prefetch("attendees", queryset=Attendee.objects.filter( Q(**attendee_filter_kwargs_master) | Q(**attendee_filter_kwargs_children_no_product) | Q(**attendee_filter_kwargs_children_product)).select_related(
            "purchase__order", "purchase__product", "contact", "contact__user"
        ).order_by(
            "purchase__order__submitted_time", "added_time"
        ), to_attr="conf_attendees")
    ).order_by("master__id")

    #events = Event.objects.filter(Q(parent__id=master_id) | Q(master__id=master_id), publish_status="PUBLISHED")

    response = HttpResponse(content_type="application/vnd.openxml")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=Event_" + report_type + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Event_" + report_type

    if events:

        row_num = 0

        columns = [
            ("Activity ID",20), 
            ("Activity Name",20), 
            #("Activity Code",20),
            ("WebUserID",10),
            ("First Name",20), 
            ("Middle Name",20), 
            ("Last Name",20), 
            ("Suffix", 10),
            #("iMIS Code",15),
            #("Registration option", 20), 
            ("Designation",20),
            ("Member Type",20),
            #("Attendee Title", 20), 
            ("Company",20), 
            ("Address 1",20), 
            ("Address 2",20), 
            ("City",20), 
            ("State",20), 
            ("Zip Code", 10),
            ("Country", 20), 
            ("Email", 20),
            ("Phone",10), 
            ("Purchase Date", 20),
            #("Quantity", 10),
            ]


        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for event in events:
            # attendee_filter_kwargs["event__id"] = event.id
            # #print("attendee filter kwargs: " + str(attendee_filter_kwargs))
            # if hasattr(event, "product"):
            #     attendees = Attendee.objects.filter(**attendee_filter_kwargs).exclude(Q(status="R") | Q(status="C")).select_related("contact").select_related("purchase").order_by("purchase__order__submitted_time")
            # else:
            #     attendees = Attendee.objects.filter(**attendee_filter_kwargs_no_product).exclude(Q(status="R") | Q(status="C")).select_related("contact").order_by("added_time")

            #print('attendees returned: ' + str(attendees))
            for attendee in event.conf_attendees:

                submitted_time = None
                purchase_quantity = None

                if hasattr(event, "product"):
                    submitted_time = attendee.purchase.order.submitted_time
                    #purchase_quantity = attendee.purchase.quantity
                else:
                    submitted_time = attendee.added_time
                    #purchase_quantity = 1

                # if row_num == 0:
                #     c1 = ws.cell(row = 1, column = 8)
                #     c2 = ws.cell(row = 1, column = 9)
                #     c3 = ws.cell(row = 1, column = 10)
            
                #     if hasattr(attendee.purchase, "product"):
                #         question_1 = attendee.purchase.product.question_1
                #         if question_1 == "":
                #             question_1 = "Question 1"
                #         c1.value = question_1
                #         question_2 = attendee.purchase.product.question_2
                #         if question_2 == "":
                #             question_2 = "Question 2"
                #         c2.value = question_2
                #         question_3 = attendee.purchase.product.question_3
                #         if question_3 == "":
                #             question_3 = "Question 3"
                #         c3.value = question_3
                #     else:
                #         c1.value = 'Question 1'
                #         c2.value = 'Question 2'
                #         c3.value = 'Question 3'

                row_num += 1
                row = [
                    event.master.id,
                    event.title,
                    #attendee.purchase.code,
                    attendee.contact.user.username,
                    attendee.contact.first_name,
                    attendee.contact.middle_name,
                    attendee.contact.last_name,
                    attendee.contact.suffix_name,
                    #purchase.product.imis_code,
                    #purchase.option,
                    attendee.contact.designation,
                    attendee.contact.member_type,
                    #contact.member_title,
                    attendee.contact.company,
                    attendee.contact.address1,
                    attendee.contact.address2,
                    attendee.contact.city,
                    attendee.contact.state,
                    attendee.contact.zip_code,
                    attendee.contact.country,
                    attendee.contact.email,
                    attendee.contact.phone,
                    submitted_time.date(),
                    # purchase_quantity, #Attendee.objects.filter(event = event, contact=attendee.contact).count() if event.event_type == "ACTIVITY" else 0,
                    ]


                for col_num in range(len(row)):
                    c = ws.cell(row=row_num+1, column=col_num+1)
                    try:
                        c.value = float(row[col_num])
                    except:
                        c.value = str(row[col_num])
        wb.save(response)
        return response
    else:
        messages.error(request, "No records were found with the dates passed in. Please change the begin and end date filters and try again.")
        return HttpResponseRedirect("/events/manage/{0}/reports/".format(str(master_id)))

def financialsummaryreport(request, master_id, report_type, begin_time, end_time):
    """
    do not require date inputs for financial reports
    """

    request_options = request.GET
    filter_kwargs = {"master__id":master_id, "publish_status":"PUBLISHED" }
    attendee_filter_kwargs = {}

    events = Event.objects.filter(Q(master_id=master_id) | Q(parent_id=master_id), publish_status='PUBLISHED')

    if events: 

        response = HttpResponse(content_type="application/vnd.openxml")
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Event_" + report_type + ".xlsx"

        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "Event_" + report_type

        row_num = 0

        columns = [
            ("Activity ID", 20),
            ("Event Name", 20),
            ("Product Amount Total", 10),
            #("Payments Total", 10),
            ("Number of purchases", 10),
            ]


        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for event in events:
            row_num += 1
            purchases = Purchase.objects.filter(product__content=event, status="A").exclude(order__isnull=True, user__isnull=True)
            #payments = Payment.objects.filter(order_id__in=[p.order.id for p in purchases])
            purchase_total = purchases.aggregate(total=Sum(F("quantity") * F("submitted_product_price_amount"))).get('total')
            #purchase_total = float("{0:.2f}".format(purchase_total))
            quantity_purchased_total = purchases.aggregate(Sum("quantity")).get("quantity__sum")

            if not purchase_total:
                purchase_total = 0.00

            if not quantity_purchased_total:
                quantity_purchased_total = 0

            row = [
                event.master.id,
                event.title,
                float("{0:.2f}".format(purchase_total)),
                #payments.aggregate(Sum('amount')).get('amount__sum'),
                quantity_purchased_total,
                ]

            for col_num in range(len(row)):
                c = ws.cell(row=row_num+1, column=col_num+1)
                try:
                    c.value = float(row[col_num])
                except:
                    c.value = str(row[col_num])

        wb.save(response)
        return response
    else:
        return messages.error(request, "Event does not exist...")


def activitysummaryreport(request, master_id, report_type, begin_time, end_time):

    request_options = request.GET
    filter_kwargs = {"master__id":master_id, "publish_status":"PUBLISHED" }

    events = Event.objects.filter(parent__id=master_id, publish_status='PUBLISHED')

    response = HttpResponse(content_type="application/vnd.openxml")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=Event_" + report_type + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Event_" + report_type

    row_num = 0

    columns = [
        ("Activity ID",20), 
        ("Title",20), 
        ("Code",20),
        ("Begin Date and Time",20), 
        ("End Date and Time", 20),
        ("Maximum Available", 20), 
        ("Tickets Sold", 20),
        ("Tickets Remaining", 20) 
        ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for event in events:
        product_max_quantity = '--'
        registered_attendees = '--'
        tickets_remaining = '--'
        if hasattr(event, "product"):
            product_max_quantity = event.product.max_quantity if hasattr(event, "product") else  '--'
            if not product_max_quantity or product_max_quantity == None:
                product_max_quantity = '--'
            registered_attendees = Attendee.objects.filter(event = event, status="A").count()
            if not registered_attendees or registered_attendees == None:
                registered_attendees = '--'

            if registered_attendees != '--' and product_max_quantity != '--':
                tickets_remaining = product_max_quantity - registered_attendees
        else:
            registered_attendees =  Attendee.objects.filter(event = event, status="A").count()


        row_num += 1
        row = [
            event.master.id,
            event.title,
            event.code,
            event.begin_time.strftime("%Y-%m-%d %H:%M:%S"),
            event.end_time.strftime("%Y-%m-%d %H:%M:%S"),
            product_max_quantity,
            registered_attendees,
            tickets_remaining
            ]

        for col_num in range(len(row)):
            c = ws.cell(row=row_num+1, column=col_num+1)
            try:
                c.value = float(row[col_num])
            except:
                c.value = str(row[col_num])

    wb.save(response)
    return response


def cancellationreport(request, master_id, report_type, begin_time, end_time):

    request_options = request.GET
    filter_kwargs = { "master__id":master_id, "publish_status":"PUBLISHED" } #

    events = Event.objects.filter(Q(master__id=master_id) | Q(parent__id=master_id), publish_status="PUBLISHED")

    response = HttpResponse(content_type="application/vnd.openxml")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=Event_" + report_type + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Event_" + report_type

    row_num = 0

    columns = [
        ("Activity ID", 20),
        ("Activity Name", 50),
        ("WebUserID", 10),
        ("First Name",20), 
        ("Middle Name",20), 
        ("Last Name",20),
        ("Suffix", 10),
        ("Designation",20),
        ("Company",20), 
        # ("Address 1",20), 
        # ("Address 2",20), 
        # ("City",20),
        # ("State",20), 
        # ("Zip Code", 10),
        # ("Country", 20), 
        ("Email", 20), 
        ("Phone",10),
        ("Price Rule Title", 20),
        ("Attendee Status", 5),]


    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]


    for event in events:

        cancelled_attendees = Attendee.objects.filter(Q(status="C") | Q(status="R"), event=event).select_related("purchase").select_related("contact", "contact__user")

        for attendee in cancelled_attendees:
            row_num += 1

            row = [
                event.master_id,
                event.title,
                attendee.contact.user.username,
                attendee.contact.first_name,
                attendee.contact.middle_name,
                attendee.contact.last_name,
                attendee.contact.suffix_name,
                attendee.contact.designation,
                attendee.contact.company,
                attendee.contact.email,
                attendee.contact.phone,
                # contact_obj.address2,
                # contact_obj.city,
                # contact_obj.state,
                # contact_obj.zip_code,
                # contact_obj.country,
                attendee.purchase.product_price.title,
                attendee.status,]

            for col_num in range(len(row)):
                c = ws.cell(row=row_num+1, column=col_num+1)
                try:
                    c.value = float(row[col_num])
                except:
                    c.value = str(row[col_num])


    wb.save(response)
    return response

def all_purchase_report(request, master_id, report_type, begin_time, end_time):

    """
    add purchase option code?
    """
    request_options = request.GET
    filter_kwargs = { "parent__id":master_id, "publish_status":"PUBLISHED" } 
    attendee_filter_kwargs = {"status":"A"}

    if begin_time and end_time:
        attendee_filter_kwargs["purchase__order__submitted_time__range"] = (begin_time, end_time)
    elif begin_time:
        attendee_filter_kwargs["purchase__order__submitted_time__gt"] = begin_time
    elif end_time:
        attendee_filter_kwargs["purchase__order__submitted_time__lt"] = end_time

    events = Event.objects.filter(Q(parent__id=master_id) | Q(master__id=master_id), publish_status="PUBLISHED").exclude(product__isnull=True).prefetch_related(
        Prefetch("attendees", queryset=Attendee.objects.filter(**attendee_filter_kwargs).exclude(purchase__isnull=True).select_related(
            "purchase__order", "contact", "contact__user"
        ).order_by(
            "purchase__order__submitted_time", "added_time"
        ), to_attr="conf_attendees")
        ).order_by("master__id")
        

    # events = Event.objects.filter(**filter_kwargs).select_related("product").prefetch_related(
    #     Prefetch("attendees", queryset=Attendee.objects.filter( Q(**attendee_filter_kwargs) | Q(**attendee_filter_kwargs_no_product), status="A").select_related(
    #         "purchase__order", "purchase__product", "contact"
    #     ).order_by(
    #         "purchase__order__submitted_time", "added_time"
    #     ), to_attr="conf_attendees")
    # ).order_by("master__id")


    response = HttpResponse(content_type="application/vnd.openxml")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=Event_" + report_type + ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Event_" + report_type


    row_num = 0

    columns = [
        ("Activity ID",20), 
        ("Activity Name",20), 
        ("Order ID",20),
        ("WebUserID", 10),
        ("First Name",20), 
        ("Middle Name",20), 
        ("Last Name",20), 
        ("Suffix", 10),
        #("iMIS Code",15),
        #("Registration option", 20), 
        ("Designation",20),
        ("Member Type",20),
        #("Attendee Title", 20), 
        ("Company",20), 
        ("Address 1",20), 
        ("Address 2",20), 
        ("City",20), 
        ("State",20), 
        ("Zip Code", 10),
        ("Country", 20), 
        ("Email", 20),
        ("Phone",10), 
        ("Purchase Date", 20),
        #("Quantity", 10),
        ("Purchase Amount", 20),
        #("Payment Method", 15),
        ("Is Manual Order", 10),
        ]


    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for event in events:
        #attendee_filter_kwargs["event__id"] = event.id
        #print('event id: ' + str(event.id))

        #print("attendee filter kwargs: " + str(attendee_filter_kwargs))
        #attendees = Attendee.objects.filter(**attendee_filter_kwargs).exclude(purchase__order__isnull=True).select_related("purchase").select_related("contact").order_by("contact")
        #print('attendees: ' + str(attendees))
        #print('attendees returned: ' + str(attendees))

        for attendee in event.conf_attendees:

            # payments = attendee.purchase.order.payment_set.all()
            # payment_type = ""

            # if payments:
            #     payment_type = ",".join(str(payment.method) for payment in payments)
            # if not payments:
            #     payment_type = "NONE"

            row_num += 1
            row = [
                event.master.id,
                event.title,
                attendee.purchase.order.id,
                attendee.contact.user.username,
                attendee.contact.first_name,
                attendee.contact.middle_name,
                attendee.contact.last_name,
                attendee.contact.suffix_name,
                attendee.contact.designation,
                attendee.contact.member_type,
                attendee.contact.company,
                attendee.contact.address1,
                attendee.contact.address2,
                attendee.contact.city,
                attendee.contact.state,
                attendee.contact.zip_code,
                attendee.contact.country,
                attendee.contact.email,
                attendee.contact.phone,
                attendee.purchase.order.submitted_time.strftime("%Y-%m-%d %H:%M:%S"),
                #attendee.purchase.quantity,
                #Attendee.objects.filter(event = event, contact=attendee.contact, status="A").count(),
                #Purchase.objects.filter(contact=attendee.contact,status="A", product__content=event).aggregate(total=Sum(F("quantity") * F("submitted_product_price_amount"))).get('total'),
                float("{0:.2f}".format(attendee.purchase.submitted_product_price_amount)),

               # payment_type,
                attendee.purchase.order.is_manual,
                ]
               
           
            for col_num in range(len(row)):
                c = ws.cell(row=row_num+1, column=col_num+1)

                try:
                    c.value = float(row[col_num])
                except:
                    c.value = str(row[col_num])

                #c.value = str(row[col_num])

    wb.save(response)
    return response

def event_option_report(request, master_id, report_type, begin_time, end_time):
    """
    do not require date inputs for financial reports
    """

    request_options = request.GET
    filter_kwargs = {"master__id":master_id, "publish_status":"PUBLISHED" }
    attendee_filter_kwargs = {}

    event = Event.objects.get(master_id=master_id, publish_status='PUBLISHED')

    if event: 

        response = HttpResponse(content_type="application/vnd.openxml")
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Event_" + report_type + ".xlsx"

        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "Event_" + report_type

        row_num = 0

        columns = [
            ("Event Option Name", 20),
            ("Number of purchases", 10),
            ]


        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for option in event.product.options.all():
            row_num += 1
            purchases = Purchase.objects.filter(product__content=event, status="A", option=option).exclude(order__isnull=True, user__isnull=True)
            #payments = Payment.objects.filter(order_id__in=[p.order.id for p in purchases])
            #urchase_total = purchases.aggregate(total=Sum(F("quantity") * F("submitted_product_price_amount"))).get('total')
            #purchase_total = float("{0:.2f}".format(purchase_total))
            quantity_purchased_total = purchases.aggregate(Sum("quantity")).get("quantity__sum")

            # if not purchase_total:
            #     purchase_total = 0.00

            if not quantity_purchased_total:
                quantity_purchased_total = 0

            row = [
                option.title,
                #float("{0:.2f}".format(purchase_total)),
                #payments.aggregate(Sum('amount')).get('amount__sum'),
                quantity_purchased_total,
                ]

            for col_num in range(len(row)):
                c = ws.cell(row=row_num+1, column=col_num+1)
                try:
                    c.value = float(row[col_num])
                except:
                    c.value = str(row[col_num])

        wb.save(response)
        return response
    else:
        return messages.error(request, "Event does not exist...")
