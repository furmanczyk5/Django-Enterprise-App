import openpyxl

from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment

from django.db.models import fields
from django.http import HttpResponse

from content.models import ContentTagType

# from events.models import NATIONAL_CONFERENCE_MASTER_ID
from conference.models import NationalConferenceActivity
from learn.models import LearnCourse, LearnCourseInfo

from events.models import EVENTS_NATIONAL_TRACK_CURRENT


def national_export_xlsx(modeladmin, request, queryset):
    if "NationalConferenceActivityAdmin" in modeladmin.__str__():
        is_npc = True
    elif "LearnCourseAdmin" in modeladmin.__str__():
        is_npc = False

    tag_type_codes = [
        "EVENTS_NATIONAL_TYPE", "ROOM", "EVENTS_NATIONAL_CEU",
        EVENTS_NATIONAL_TRACK_CURRENT, "NPC_TOPIC", "EVENTS_NATIONAL_AUDIENCE",
        "DIVISION", "TRANSIT"]

    ignore_fields = [
        "id", "publish_uuid", "created_by", "updated_by", "master",
        "content_type", "publish_status", "event_type", "has_xhtml",
        "publish_time", "archive_time", "parent", "abstract", "resource_type",
        "serial_pub", "volume_number", "issue_number", "language",
        "resource_published_date", "copyright_date", "copyright_statement",
        "isbn", "content_ptr", "user_address_num", "address1", "address2",
        "city", "state", "zip_code", "country"]

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if is_npc:
        response['Content-Disposition'] = 'attachment; filename=national_activities_export.xlsx'
    else:
        response['Content-Disposition'] = 'attachment; filename=learn_course_export.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    if is_npc:
        ws.title = "National Conference Activities"
        activity_fields = NationalConferenceActivity._meta.fields
        all_fields = activity_fields
    else:
        ws.title = "APA Learn Courses"
        activity_fields = LearnCourse._meta.fields
        activity_fields = list(activity_fields)
        fl = list(LearnCourseInfo._meta.fields)
        vimeo_field = [f for f in fl if f.name is 'vimeo_id']
        all_fields = activity_fields + vimeo_field

    content_tag_types = ContentTagType.objects.filter(
        # content__parent__id=NATIONAL_CONFERENCE_MASTER_ID
        content__in=queryset
        ).select_related("content"
        ).select_related("tag_type"
        ).prefetch_related("tags")

    # NOT GREAT... struggling with prefetch_related to pull
    # tag_types / tags through the existing queryset,
    # so instead just pulling ALL tags
    # as a separate query, and then adding to a dictionary...
    tags_all = {}
    for ct in content_tag_types:
    # for ct in ContentTagType.objects.filter(content__parent__id = NATIONAL_CONFERENCE_MASTER_ID).select_related("content").select_related("tag_type").prefetch_related("tags"):
        if ct.content.id not in tags_all:
            tags_all[ct.content.id] = {}
        tags_list = [t.title for t in ct.tags.all()]
        tags_all[ct.content.id][ct.tag_type.code] = ",".join(tags_list)

    header_style = NamedStyle(name='header_style', font=Font(bold=True))
    value_style = NamedStyle(name='value_style', alignment=Alignment(wrap_text=True))

    activities = queryset.prefetch_related("master")

    # activities = queryset.prefetch_related(
    #     Prefetch('tag_types',
    #         queryset=Content.objects.all().select_related('tag_types'))
    #     ).all()
    # a = activities.first()
    # print(a.contenttagtype.first().tag_type.title)

    # add master first as a special case:
    c = ws.cell(row=1, column=1)
    c.value = "Master ID"
    c.style = header_style

    row_num = 1
    col_num = 2

    # add the fields associated with the activity record
    for field in all_fields:
        if field.name not in ignore_fields:
            c = ws.cell(row = row_num, column= col_num)
            c.value = field.verbose_name
            c.style = header_style
            # setting specific column widths in a few cases:
            if field.name == "text":
                ws.column_dimensions[get_column_letter(col_num)].width = 120
            elif field.name == "description" or field.name == "title" :
                ws.column_dimensions[get_column_letter(col_num)].width = 60
            else:
                ws.column_dimensions[get_column_letter(col_num)].width = 22

            col_num += 1

    # add the tags:
    for tag_type_code in tag_type_codes:
        c = ws.cell(row=row_num, column=col_num)
        c.value = tag_type_code
        c.style = header_style
        col_num += 1

    # now add all the values...
    row_num = 2

    for activity in activities:

        # add master id value first as special case
        c = ws.cell(row=row_num, column=1)
        c.value = activity.master.id
        c.style = header_style

        col_num = 2

        # add activity record values
        for field in all_fields:
            if field.name not in ignore_fields:
                c = ws.cell(row=row_num, column=col_num)
                if isinstance(field, fields.DateTimeField):
                    if field.name == 'begin_time':
                        mytime = activity.begin_time_astimezone()
                        if mytime is not None:
                            c.value = mytime.strftime("%Y-%m-%d %I:%M %p")
                    elif field.name == 'end_time':
                        mytime = activity.end_time_astimezone()
                        if mytime is not None:
                            c.value = mytime.strftime("%Y-%m-%d %I:%M %p")
                    else:
                        mytime = getattr(activity, field.name)
                        if mytime is not None:
                            c.value = mytime.strftime("%Y-%m-%d %I:%M %p")
                elif field.name is 'vimeo_id':
                    lci = LearnCourseInfo.objects.filter(
                        learncourse=activity).first()
                    c.value = str(getattr(lci, field.name)) if lci else None
                elif not isinstance(field, fields.related.ForeignKey):
                    # just in case there is some crazy situation
                    c.value = str(getattr(activity, field.name))
                c.style = value_style
                col_num += 1

        # add tag values:
        for tag_type_code in tag_type_codes:
            if activity.id in tags_all:
                ttc = tag_type_code
                if any(ttc in tt_code for tt_code in tags_all[activity.id]):
                # if tag_type_code in tags_all[activity.id]:
                    c = ws.cell(row=row_num, column=col_num)
                    val_list = [v for k, v in tags_all[activity.id].items() if ttc in k]
                    c.value = val_list[0] if val_list else None
                    # c.value = tags_all[activity.id][tag_type_code]

            # this is what we were trying with the existing queryset... created LOTS of extra db queries
            # try:
            #     # activity_contenttagtype = contenttagtypes.get(content=activity, tag_type__code=tag_type_code)
            #     # if activity_contenttagtype is not None:
            #         # tags_list = [t.title for t in activity_contenttagtype.tags.all()]
            #         # pass
            #     # tag_types = activity.contenttagtype
            #     # tag_type = activity.contenttagtype.get(tag_type__code=tag_type_code)
            #     # tags_list = [t.title for t in tag_type.tags.all()]
            #     # c.value = ",".join(tags_list)
            #     pass
            # except (ContentTagType.DoesNotExist, TagType.DoesNotExist):
            #     pass

            col_num += 1

        row_num += 1

    wb.save(response)
    return response
national_export_xlsx.short_description = "Export XLSX"
